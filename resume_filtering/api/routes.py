import os
import io
from dotenv import load_dotenv

import tempfile
from typing import Optional
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
import uuid
from pymongo import MongoClient
from resume_filtering.services.resume_Parser import extract_text_from_pdf
from resume_filtering.services.segmentation import segment_sections
from datetime import datetime
from resume_filtering.services.informationExtractor import (
    nameExtractor, email_extractor, phone_extractor,
    skillExtractor, experienceExtractor, educationExtractor,
    certificationExtractor, languageExtractor, projectExtractor
)
from resume_filtering.services.normalizeSKillsPipeline import normalize_skills_pipeline
from resume_filtering.services.structuredExperience import extract_experience
from resume_filtering.services.structuredEducation import structure_education
from resume_filtering.services.jobpipeline import job_pipeline
from resume_filtering.services.similarityChecking import (
    calculateSimilarity, calculateExperienceSimilarity,
    weightedSimilarity
)





load_dotenv()
uri= os.getenv("MONGO-KEY")

client= MongoClient(uri)
db = client.resume_reviewer
sessionsCollection= db.sessions


#for auto deleting document data after 24 hours
sessionsCollection.create_index("created_at",expireAfterSeconds=86400)


router = APIRouter()

@router.get("/healthz")
def root():
    return {"status": "online", "message": "AI Resume Filter API"}

@router.post("/filter")
async def filter_resume(
    resumes: list[UploadFile] = File(..., description="Candidate's resumes in PDF format"),
    job_skills: str = Form(..., description="Required skills, comma-separated"),
    required_experience: Optional[str] = Form(None, description="Experience requirements"),
):
    if len(resumes) > 30:
        raise HTTPException(status_code=400, detail="Max 30 resumes allowed")

    screen_summaries = []
    resultStore=[]
    for resume in resumes:
        if not resume.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail=f"{resume.filename} is not a PDF.")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            content = await resume.read()
            tmp.write(content)
            tmp_path = tmp.name

        try:
            documents = extract_text_from_pdf(tmp_path, resume.filename)
            if not documents:
                raise HTTPException(status_code=422, detail="Could not extract text.")

            full_text = ""
            segments = {}
            for doc in documents:
                full_text += doc.page_content + "\n"
                segments.update(segment_sections(doc.page_content))

            name            = nameExtractor(full_text)
            email           = email_extractor(full_text)
            phone           = phone_extractor(full_text)
            skills_raw      = skillExtractor(segments)
            experience_raw  = experienceExtractor(segments)
            education_raw   = educationExtractor(segments)
            certifications  = certificationExtractor(segments)
            languages       = languageExtractor(segments)
            projects        = projectExtractor(segments)

            normalized_skills      = normalize_skills_pipeline(skills_raw)
            structured_experience  = extract_experience(experience_raw)
            structured_education   = structure_education(education_raw)
            job_phrases, role      = job_pipeline(job_skills, required_experience or "") # here role -> list[dict{[roles],months},...] of job description
            # jbSet = set([skill.strip().lower() for skill in job_phrases])
            # candSet = set([skill.strip().lower() for skill in normalized_skills])
            # matched_skills= list(jbSet.intersection(candSet))

            skill_score      = calculateSimilarity(job_phrases, normalized_skills)
            exp_score        = calculateExperienceSimilarity(role, structured_experience)
            weighted_score   = weightedSimilarity(job_phrases, normalized_skills, role, structured_experience)

            #handles diff experiences seprately
            experience_columns={}
            for index, job in enumerate(structured_experience):
                job_number = index + 1 # Starts at 1 instead of 0
                experience_columns[f"Exp {job_number} Role"] = job.get("titles") or "Unknown"
                companies_list = job.get("companies", [])
                experience_columns[f"Exp {job_number} Company"] = ", ".join(companies_list) if companies_list else "Unknown"
                experience_columns[f"Exp {job_number} Duration"] = job.get("duration_months") or 0
            total_months = sum([job.get("duration_months", 0) for job in structured_experience if job.get("duration_months")])

            full_record = {
                "name":  name or "Unknown",
                "email": email or "",
                "phone": phone or "",
                "skill_score(%)":      round(skill_score, 3),
                "experience_score(%)": round(exp_score, 3),
                "weighted_score(%)":   round(weighted_score, 3),
                "skills":          normalized_skills,
                "total_experience_months": total_months,
                "certifications":  certifications,
                "experience":      structured_experience, #  for JSON
                # "experience_roles": exp_roles,            # For Excel
                # "companies":       list(set(all_companies)),
                **experience_columns,
                "education":       structured_education,
                "languages":       languages,
                "projects":        projects,
                # "matched_skills": matched_skills,

                "filename":            resume.filename,
                "processed_at":        datetime.now().isoformat(timespec="seconds"),
                "job_skills":          job_skills,
                "required_experience": required_experience,
            }

            resultStore.append(full_record)
            screen_summaries.append({
                "name":             full_record["name"],
                "email":            full_record["email"],
                # "matchedSkills":    full_record["matched_skills"],
                "skill_score":      full_record["skill_score(%)"],
                "experience_score":      full_record["experience_score(%)"],
                "weighted_score":   full_record["weighted_score(%)"],

            })

        finally:
            try:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
            except PermissionError:
                pass
    screen_summaries.sort(key=lambda x :x ["weighted_score"],reverse=True)
    resultStore.sort(key=lambda x :x ["weighted_score(%)"],reverse=True)
    sessionID=str(uuid.uuid4())
    document = {
        "session_id": sessionID,
        "created_at": datetime.utcnow(),
        "results": resultStore
    }
    sessionsCollection.insert_one(document)
    return {
        "message": f"Successfully processed {len(resumes)} resumes.",
        "session_id": sessionID,
        "results": screen_summaries
    }


def build_excel(data_store):
    """
    Creates an Excel file in memory from the resultStore and returns raw bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resume Comparison"
    longestRecord= max(data_store,key=len)
    headers= list(longestRecord.keys())
    ws.append(headers)

    # Sort data by highest weighted score before exporting
    sorted_data= sorted(data_store,key= lambda x : x["weighted_score(%)"],  reverse= True)
    for row in sorted_data:
        rowData=[]
        for col in headers:
            val = row.get(col,"")
            if col in ["skill_score(%)", "experience_score(%)", "weighted_score(%)"]:
                if isinstance(val, (int, float)):
                    val = val*100.0
            if isinstance(val, (list,dict)):
                rowData.append(str(val))
            else :
                rowData.append(val)
        ws.append(rowData)
    headerFill= PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    headerFont= Font(color="FFFFFF", bold=True)


    for cell in ws[1]: #sets header
        cell.fill=headerFill
        cell.font=headerFont
        cell.alignment=Alignment(horizontal="left",vertical="center")


    for colIndex,colCells in enumerate(ws.columns,start=1): #sets column width for clear reading
        maxLen = 0
        for col in colCells:
            if col.value:
                colLen=len(str(col.value))
                if colLen > maxLen:
                    maxLen=colLen


        adjustedWidth = min (maxLen + 3,80)
        colLetter= get_column_letter(colIndex)
        ws.column_dimensions[colLetter].width=adjustedWidth

    virtual_wb= io.BytesIO()
    wb.save(virtual_wb)
    virtual_wb.seek(0)
    return virtual_wb.read()

@router.delete("/results/{session_id}", tags=["Delete"])
def clear_results(session_id : str):
    result = sessionsCollection.delete_one({"session_id": session_id})
    if result.deleted_count == 1:
        return {"message": "Cleared. Ready for the new files"}
    else:
        raise HTTPException(status_code=404, detail="couldn't clear results.")


@router.get("/export/{session_id}", tags=["Export"])
def export_excel(session_id:str):
    sessionData= sessionsCollection.find_one({"session_id":session_id})
    if not sessionData:
        raise HTTPException(
            status_code=404,
            detail="Session expired or not found. Please run the analysis again."
        )
    cleanResult=sessionData["results"]
    excel_bytes = build_excel(cleanResult)

    filename = f"resume_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )